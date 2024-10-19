from tkinter import * #importing all tkinter widgets
from tkinter import messagebox # Importing the messagebox module from tkinter for displaying pop-up messages.
import random # Importing the random module for generating random numbers.

# Importing the os module for interacting with the operating system.
import os 
import random, os, tempfile, smtplib
import openpyxl

# Create a new Excel workbook
wb = openpyxl.load_workbook('customer_details.xlsx')
sheet = wb.active

def store_customer_details():
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Phone Number'
    sheet['C1'] = 'Bill No.'

    try:
     if nameEntry.get()=='' or phoneEntry.get()=='':
        messagebox.showerror('Error','Customer Details Are Required')
     elif cosmeticpriceEntry.get()=='' and grocerypriceEntry.get()=='' or statpriceEntry.get()=='':
        messagebox.showerror('Error', 'No Products are selected') 
     elif cosmeticpriceEntry.get()=='0 ₹' and grocerypriceEntry.get()=='0 ₹' and statpriceEntry.get=='0 ₹':
        messagebox.showerror('Error', 'No Products are selected')

     else:
        sheet.cell(row=sheet.max_row + 1, column=1, value=nameEntry.get())
        sheet.cell(row=sheet.max_row, column=2, value=phoneEntry.get())
        sheet.cell(row=sheet.max_row, column=3, value=billnumber)
        wb.save('customer_details.xlsx')
        messagebox.showinfo('Saved','Customer Details Saved Successfully')

    except Exception as e:
        messagebox.showerror('Error', 'Close the Excel sheet and Try again'.format(str(e)))


# Functionality Part

def clear():
    bathsoapEntry.delete(0,END)
    facecreamEntry.delete(0,END)
    facewashEntry.delete(0,END)
    shampooEntry.delete(0,END)
    bodylotionEntry.delete(0,END)
    bodywashEntry.delete(0,END)

    riceEntry.delete(0,END)
    oilEntry.delete(0,END)
    dalEntry.delete(0,END)
    wheatEntry.delete(0,END)
    sugarEntry.delete(0,END)
    teaEntry.delete(0,END)

    envelopeEntry.delete(0,END)
    penEntry.delete(0,END)
    calciEntry.delete(0,END)
    notebookEntry.delete(0,END)
    staplerEntry.delete(0,END)
    glueEntry.delete(0,END)

    bathsoapEntry.insert(0,0)
    facecreamEntry.insert(0,0)
    facewashEntry.insert(0,0)
    shampooEntry.insert(0,0)
    bodylotionEntry.insert(0,0)
    bodywashEntry.insert(0,0)

    riceEntry.insert(0,0)
    oilEntry.insert(0,0)
    dalEntry.insert(0,0)
    wheatEntry.insert(0,0)
    sugarEntry.insert(0,0)
    teaEntry.insert(0,0)

    envelopeEntry.insert(0,0)
    penEntry.insert(0,0)
    calciEntry.insert(0,0)
    notebookEntry.insert(0,0)
    staplerEntry.insert(0,0)
    glueEntry.insert(0,0)

    cosmetictaxEntry.delete(0,END)
    grocerytaxEntry.delete(0,END)
    stattaxEntry.delete(0,END)

    cosmeticpriceEntry.delete(0,END)
    grocerypriceEntry.delete(0,END)
    statpriceEntry.delete(0,END)

    nameEntry.delete(0,END)
    phoneEntry.delete(0,END)
    textarea.delete('1.0',END)
    billnumberEntry.delete(0,END)

def send_email():
    def send_gmail():
        try:
            ob=smtplib.SMTP('smtp.gmail.com', 587)
            ob.starttls()
            ob.login(senderEntry.get(), passwordEntry.get())
            message=email_textarea.get(1.0, END)
            ob.sendmail(senderEntry.get(), recieverEntry.get(), message.encode('utf-8'))
            ob.quit()
            messagebox.showinfo('Success', 'Bill is successfully sent', parent=root1)
            root1.destroy()
        
        except smtplib.SMTPAuthenticationError as e:
            messagebox.showerror('Error', 'Authentication error: ' + str(e), parent=root1)
        except smtplib.SMTPException as e:
            messagebox.showerror('Error', 'SMTP error: ' + str(e), parent=root1)
        except Exception as e:
            messagebox.showerror('Error', 'An error occurred: ' + str(e))
    if textarea.get(1.0, END) == '\n':
        messagebox.showerror('Error', 'Bill is empty')
    else:
        root1= Toplevel()
        root1.grab_set()
        root1.title('Email')
        root1.config(bg='gray20')
        root1.resizable(0,0)

        senderFrame=LabelFrame(root1, text='SENDER', font=('arial', 14, 'bold'), bd=6, bg='gray20', fg='white', relief=GROOVE)
        senderFrame.grid(row=0, column=0, padx=40, pady=20)

        senderLabel=Label(senderFrame, text="Sender's Email", font=('arial', 14, 'bold'), bg='gray20', fg='white')
        senderLabel.grid(row=0, column=0, padx=10, pady=8)

        senderEntry=Entry(senderFrame, font=('arial', 14, 'bold'), bd=5, width=23, relief=RIDGE)
        senderEntry.grid(row=0, column=1, padx=10, pady=8)
        
        #Email: ketankurve16@gmail.com
        #Password: zanl qoht ybft zmyf
        passwordLabel=Label(senderFrame, text="Password", font=('arial', 14, 'bold'), bg='gray20', fg='white')
        passwordLabel.grid(row=1, column=0, padx=10, pady=8)

        passwordEntry=Entry(senderFrame, font=('arial', 14, 'bold'), bd=5, width=23, relief=RIDGE, show='*')
        passwordEntry.grid(row=1, column=1, padx=10, pady=8)

        recipientFrame=LabelFrame(root1, text='RECIPIENT', font=('arial', 14, 'bold'), bd=6, bg='gray20', fg='white', relief=GROOVE)
        recipientFrame.grid(row=1, column=0, padx=40, pady=20)

        recieverLabel=Label(recipientFrame, text="Reciever's Email", font=('arial', 14, 'bold'), bg='gray20', fg='white')
        recieverLabel.grid(row=0, column=0, padx=10, pady=8)

        recieverEntry=Entry(recipientFrame, font=('arial', 14, 'bold'), bd=5, width=23, relief=RIDGE)
        recieverEntry.grid(row=0, column=1, padx=10, pady=8)

        messageLabel=Label(recipientFrame, text="Message", font=('arial', 14, 'bold'), bg='gray20', fg='white')
        messageLabel.grid(row=1, column=0, padx=10, pady=8)

        email_textarea=Text(recipientFrame, font=('arial', 12, 'bold'), bd=2, width=42, height=11, relief=SUNKEN)
        email_textarea.grid(row=2, column=0, columnspan=2)
        email_textarea.delete(1.0, END)
        email_textarea.insert(END, textarea.get(1.0, END).replace('=','').replace('-','').replace('\t\t\t','\t\t').replace('\t\t**Welcome Customer**','\t    **Welcome Customer**\n'))

        sendButton=Button(root1, text='SEND', font=('arial', 16, 'bold'), bd=7, width=15, cursor='hand2', command=send_gmail)
        sendButton.grid(row=2, column=0, pady=20)

        root1.mainloop()
        


def print_bill():
    if textarea.get(1.0, END) == '\n':
        messagebox.showerror('Error', 'Bill is empty')
    else:
        file = tempfile.mktemp('.txt')
        with open(file, 'w', encoding='utf-8') as f:
            f.write(textarea.get(1.0, END))
        os.startfile(file, 'print')


def search_bill():
    for i in os.listdir('bills/'):
        if i.split('.')[0] == billnumberEntry.get():
            f=open(f'bills/{i}', 'r')
            textarea.delete('1.0', END)
            for data in f:
                textarea.insert(END, data)
            f.close()
            break
    else:
        messagebox.showerror('Error', 'Invalid Bill Number')


# Checking if the directory 'bills' exists.
# If it does not exist, create it.
if not os.path.exists('bills'):
    os.mkdir('bills')

# Defining the save_bill function.
# This function saves the bill to a text file.

billnumber=random.randint(500,1000)

def save_bill():
    
    global billnumber  # Using the global keyword to access the global billnumber variable.
    # Displaying a pop-up message asking the user if they want to save the bill.
    result = messagebox.askyesno('Confirm', 'Do you want to save bill?')
    if result:
        
        bill_content = textarea.get(1.0, END) # Getting the content of the textarea.
        file = open(f'bills/{billnumber}.txt', 'w', encoding='utf-8') # Opening a file in write mode to save the bill content.
        file.write(bill_content) # Writing the bill content to the file.
        file.close()
        messagebox.showinfo('Saved',f'Bill no. {billnumber} saved successfully')
        billnumber=random.randint(500,1000) # Generating a new random number for the next bill.
def bill_area():
    if nameEntry.get()=='' or phoneEntry.get()=='':
        messagebox.showerror('Error','Customer Details Are Required')
    elif cosmeticpriceEntry.get()=='' and grocerypriceEntry.get()=='' or statpriceEntry.get()=='':
        messagebox.showerror('Error', 'No Products are selected') 
    elif cosmeticpriceEntry.get()=='0 ₹' and grocerypriceEntry.get()=='0 ₹' and statpriceEntry.get=='0 ₹':
        messagebox.showerror('Error', 'No Products are selected')

    else: 
        textarea.delete('1.0',END)
        textarea.insert(END, '\t\t**WELCOME CUSTOMER**\n')
        textarea.insert(END, f'\nBill Number: {billnumber}\n')  
        textarea.insert(END, f'\nCustomer Name: {nameEntry.get()}\n') 
        textarea.insert(END, f'\nCustomer Phone Number: {phoneEntry.get()}\n') 
        textarea.insert(END, '\n=======================================================') 
        textarea.insert(END, 'Products\t\t\tQTY\t\t\tPrice')
        textarea.insert(END, '\n=======================================================') 
        if bathsoapEntry.get()!='0':
            textarea.insert(END,f'\nBath Soap\t\t\t{bathsoapEntry.get()}\t\t\t₹ {soapprice}')
        if facecreamEntry.get()!='0':
            textarea.insert(END,f'\nFace Cream\t\t\t{facecreamEntry.get()}\t\t\t₹ {facecreamprice}')
        if facewashEntry.get()!='0':
            textarea.insert(END,f'\nFace Wash\t\t\t{facewashEntry.get()}\t\t\t₹ {facewashprice}')
        if shampooEntry.get()!='0':
            textarea.insert(END,f'\nShampoo\t\t\t{shampooEntry.get()}\t\t\t₹ {shampooprice}')
        if bodylotionEntry.get()!='0':
            textarea.insert(END,f'\nBody Lotion\t\t\t{bodylotionEntry.get()}\t\t\t₹ {bodylotionprice}')
        if bodywashEntry.get()!='0':
            textarea.insert(END,f'\nBody Wash\t\t\t{bodywashEntry.get()}\t\t\t₹ {bodywashprice}')
        
        if riceEntry.get()!='0':
            textarea.insert(END,f'\nRice\t\t\t{riceEntry.get()}\t\t\t₹ {riceprice}')
        if oilEntry.get()!='0':
            textarea.insert(END,f'\nOil\t\t\t{oilEntry.get()}\t\t\t₹ {oilprice}')
        if dalEntry.get()!='0':
            textarea.insert(END,f'\nDal\t\t\t{dalEntry.get()}\t\t\t₹ {dalprice}')
        if wheatEntry.get()!='0':
            textarea.insert(END,f'\nWheat\t\t\t{wheatEntry.get()}\t\t\t₹ {wheatprice}')
        if sugarEntry.get()!='0':
            textarea.insert(END,f'\nSugar\t\t\t{sugarEntry.get()}\t\t\t₹ {sugarprice}')
        if teaEntry.get()!='0':
            textarea.insert(END,f'\nTea\t\t\t{teaEntry.get()}\t\t\t₹ {teaprice}')

        if envelopeEntry.get()!='0':
            textarea.insert(END,f'\nEnvelopes\t\t\t{envelopeEntry.get()}\t\t\t₹ {envelopeprice}')
        if penEntry.get()!='0':
            textarea.insert(END,f'\nPens\t\t\t{penEntry.get()}\t\t\t₹ {penprice}')
        if calciEntry.get()!='0':
            textarea.insert(END,f'\nCalculators\t\t\t{calciEntry.get()}\t\t\t₹ {calciprice}')
        if notebookEntry.get()!='0':
            textarea.insert(END,f'\nNotebooks\t\t\t{notebookEntry.get()}\t\t\t₹ {notebookprice}')
        if staplerEntry.get()!='0':
            textarea.insert(END,f'\nStaplers\t\t\t{staplerEntry.get()}\t\t\t₹ {staplerprice}')
        if glueEntry.get()!='0':
            textarea.insert(END,f'\nGlue\t\t\t{glueEntry.get()}\t\t\t₹ {glueprice}')
        textarea.insert(END, '\n-------------------------------------------------------')

        if cosmetictaxEntry.get()!='₹ 0.0':
            textarea.insert(END, f'\nCosmetic Tax\t\t\t\t{cosmetictaxEntry.get()}')
        if grocerytaxEntry.get()!='₹ 0.0':
            textarea.insert(END, f'\nGrocery Tax\t\t\t\t{grocerytaxEntry.get()}')
        if stattaxEntry.get()!='₹ 0.0':
            textarea.insert(END, f'\nStationary Tax\t\t\t\t{statpriceEntry.get()}')

        textarea.insert(END, f'\n\nTotal Bill: \t\t\t\t₹ {totalbill}')
        textarea.insert(END, '\n-------------------------------------------------------')

        save_bill()
              
def total():
    global soapprice, facecreamprice, facewashprice, shampooprice, bodylotionprice, bodywashprice
    global riceprice, oilprice, dalprice, wheatprice, sugarprice, teaprice
    global envelopeprice, penprice, calciprice, notebookprice, staplerprice, glueprice
    global totalbill

    #cosmetics total calculation
    soapprice=int(bathsoapEntry.get())*20
    facecreamprice=int(facecreamEntry.get())*100
    facewashprice=int(facewashEntry.get())*100
    shampooprice=int(shampooEntry.get())*200
    bodylotionprice=int(bodylotionEntry.get())*150
    bodywashprice=int(bodywashEntry.get())*200

    totalcosmeticprice = soapprice + facecreamprice + facewashprice + shampooprice + bodylotionprice + bodywashprice
    cosmeticpriceEntry.delete(0,END)
    cosmeticpriceEntry.insert(0,'₹ '+str(totalcosmeticprice))

    #Cosmetic Tax
    cosmetictax=totalcosmeticprice*0.12
    cosmetictaxEntry.delete(0,END)# Clear the previous value in the cosmetictaxEntry widget
    cosmetictaxEntry.insert(0,'₹ '+str(cosmetictax)) # Insert the new calculated tax value into the cosmetictaxEntry widget

    #grocery total calculation
    riceprice=int(riceEntry.get())*60
    oilprice=int(oilEntry.get())*120
    dalprice=int(dalEntry.get())*150
    wheatprice=int(wheatEntry.get())*40
    sugarprice=int(sugarEntry.get())*40
    teaprice=int(teaEntry.get())*300

    totalgroceryprice=riceprice + oilprice + dalprice + wheatprice + sugarprice + teaprice
    grocerypriceEntry.delete(0,END)
    grocerypriceEntry.insert(0,'₹ '+str(totalgroceryprice))

    #Grocery Tax
    grocerytax = totalgroceryprice*0.05
    grocerytaxEntry.delete(0,END)# Clear the previous value in the grocerytaxEntry widget
    grocerytaxEntry.insert(0,'₹ '+str(grocerytax)) # Insert the new calculated tax value into the grocerytaxEntry widget

    #stationary total calculation
    envelopeprice=int(envelopeEntry.get())*10
    penprice=int(penEntry.get())*10
    calciprice=int(calciEntry.get())*1000
    notebookprice=int(notebookEntry.get())*100
    staplerprice=int(staplerEntry.get())*100
    glueprice=int(glueEntry.get())*10

    totalstatprice = envelopeprice + penprice + calciprice + notebookprice + staplerprice + glueprice
    statpriceEntry.delete(0,END)
    statpriceEntry.insert(0,'₹ '+str(totalstatprice))
        
    #Stationary Tax
    stattax = totalstatprice*0.15
    stattaxEntry.delete(0, END) # Clear the previous value in the stattaxEntry widget
    stattaxEntry.insert(0,'₹ '+str(stattax)) # Insert the new calculated tax value into the stattaxEntry widget

    totalbill=totalcosmeticprice+totalgroceryprice+totalstatprice+cosmetictax+grocerytax+stattax

#GUI Part
root = Tk()
root.title('Online Billing System')
root.geometry('1920x1080')
root.iconbitmap('icon.ico')
headingLabel=Label(root, text='Online Billing System', font=('times new roman',30,'bold'), bg='gray20', fg='gold', bd=12, relief=GROOVE)
headingLabel.pack(fill=X)
root.configure(bg='gray20') # Change the background color of the root window to gray20

#-----------------------------------------------------------------------------------------------------------------------

customer_details_frame=LabelFrame(root, text='Customer Details',font=('times new roman',15,'bold'), fg='gold', bd=8, relief=GROOVE, bg='gray20')
customer_details_frame.pack(fill=X, pady=1)

nameLabel=Label(customer_details_frame,text='Name', font=('times new roman',15,'bold'), bg='gray20', fg='white')
nameLabel.grid(row=0,column=0, padx=20)

nameEntry=Entry(customer_details_frame, font=('arial',15), bd=7, width=30)
nameEntry.grid(row=0, column=1, padx=8)

phoneLabel=Label(customer_details_frame,text='Phone Number', font=('times new roman',15,'bold'), bg='gray20', fg='white')
phoneLabel.grid(row=0,column=2, padx=20, pady=2)

phoneEntry=Entry(customer_details_frame, font=('arial',15), bd=7, width=20, validate='key', validatecommand=(root.register(lambda P: P.isdigit() and len(P) <= 10 or P == ''), '%P'))
phoneEntry.grid(row=0, column=3, padx=8)


billnumberLabel=Label(customer_details_frame,text='Bill Number', font=('times new roman',15,'bold'), bg='gray20', fg='white')
billnumberLabel.grid(row=0,column=4, padx=20, pady=2)

billnumberEntry=Entry(customer_details_frame, font=('arial',15), bd=7, width=25)
billnumberEntry.grid(row=0, column=5, padx=8)

searchButton=Button(customer_details_frame, text='SEARCH', font=('arial',12,'bold'), bd=7, width=10, cursor='hand2', command=search_bill)
searchButton.grid(row=0, column=6, padx=20, pady=8)

#-----------------------------------------------------------------------------------------------------------------------

# Product Frame
productFrame=Frame(root, bg='gray20', relief=GROOVE, bd=5)
productFrame.pack(fill=X)

# Cosmetics Frame
cosmeticsFrame=LabelFrame(productFrame, text='Cosmetics', font=('times new roman',15,'bold'), fg='gold', bd=8, relief=GROOVE, bg='gray20')
cosmeticsFrame.grid(row=0, column=0, padx=10, sticky="nsew")


def clear_initial_zero(event):
    # Function to clear the initial "0" when the user clicks on the entry field
    widget = event.widget
    if widget.get() == "0":
        widget.delete(0, "end")

def restore_initial_zero(event):
    # Function to restore the initial "0" when the user clicks outside the entry field
    widget = event.widget
    if not widget.get():
        widget.insert(0, "0")

bathsoapLabel=Label(cosmeticsFrame, text='Bath Soap  (₹20)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
bathsoapLabel.grid(row=0, column=0, pady=9, padx=10, sticky='w')
    
bathsoapEntry=Entry(cosmeticsFrame, font=('times new roman',15,'bold'), width=7, bd=5)
bathsoapEntry.grid(row=0, column=1, pady=9, padx=10)
bathsoapEntry.insert(0, 0)
bathsoapEntry.bind("<FocusIn>", clear_initial_zero)
bathsoapEntry.bind("<FocusOut>", restore_initial_zero)

facecreamLabel=Label(cosmeticsFrame, text='Face Cream  (₹100)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
facecreamLabel.grid(row=1, column=0, pady=9, padx=10, sticky='w')

facecreamEntry=Entry(cosmeticsFrame, font=('times new roman',15,'bold'), width=7, bd=5)
facecreamEntry.grid(row=1, column=1, pady=9, padx=10)
facecreamEntry.insert(0, 0)
facecreamEntry.bind("<FocusIn>", clear_initial_zero)
facecreamEntry.bind("<FocusOut>", restore_initial_zero)

facewashLabel=Label(cosmeticsFrame, text='Face Wash  (₹100)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
facewashLabel.grid(row=2, column=0, pady=9, padx=10, sticky='w')

facewashEntry=Entry(cosmeticsFrame, font=('times new roman',15,'bold'), width=7, bd=5)
facewashEntry.grid(row=2, column=1, pady=9, padx=10)
facewashEntry.insert(0, 0)
facewashEntry.bind("<FocusIn>", clear_initial_zero)
facewashEntry.bind("<FocusOut>", restore_initial_zero)

shampooLabel=Label(cosmeticsFrame, text='Shampoo  (₹200)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
shampooLabel.grid(row=3, column=0, pady=9, padx=10, sticky='w')

shampooEntry=Entry(cosmeticsFrame, font=('times new roman',15,'bold'), width=7, bd=5)
shampooEntry.grid(row=3, column=1, pady=9, padx=10)
shampooEntry.insert(0, 0)
shampooEntry.bind("<FocusIn>", clear_initial_zero)
shampooEntry.bind("<FocusOut>", restore_initial_zero)

bodylotionLabel=Label(cosmeticsFrame, text='Body Lotion  (₹150)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
bodylotionLabel.grid(row=4, column=0, pady=9, padx=10)

bodylotionEntry=Entry(cosmeticsFrame, font=('times new roman',15,'bold'), width=7, bd=5)
bodylotionEntry.grid(row=4, column=1, pady=9, padx=10)
bodylotionEntry.insert(0, 0)
bodylotionEntry.bind("<FocusIn>", clear_initial_zero)
bodylotionEntry.bind("<FocusOut>", restore_initial_zero)

bodywashLabel=Label(cosmeticsFrame, text='Body Wash  (₹200)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
bodywashLabel.grid(row=5, column=0, pady=9, padx=10)

bodywashEntry=Entry(cosmeticsFrame, font=('times new roman',15,'bold'), width=7, bd=5)
bodywashEntry.grid(row=5, column=1, pady=9, padx=10)
bodywashEntry.insert(0, 0)
bodywashEntry.bind("<FocusIn>", clear_initial_zero)
bodywashEntry.bind("<FocusOut>", restore_initial_zero)

#-----------------------------------------------------------------------------------------------------------------------
# Grocery Part
# Grocery Frame
groceryFrame=LabelFrame(productFrame, text='Grocery', font=('times new roman',15,'bold'), fg='gold', bd=8, relief=GROOVE, bg='gray20')
groceryFrame.grid(row=0, column=1, sticky="nsew")

riceLabel=Label(groceryFrame, text='Rice  (₹60)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
riceLabel.grid(row=0, column=0, pady=9, padx=10, sticky='w')

riceEntry=Entry(groceryFrame, font=('times new roman',15,'bold'), width=7, bd=5)
riceEntry.grid(row=0, column=1, pady=9, padx=10)
riceEntry.insert(0, 0)
riceEntry.bind("<FocusIn>", clear_initial_zero)
riceEntry.bind("<FocusOut>", restore_initial_zero)

oilLabel=Label(groceryFrame, text='Oil  (₹120)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
oilLabel.grid(row=1, column=0, pady=9, padx=10, sticky='w')

oilEntry=Entry(groceryFrame, font=('times new roman',15,'bold'), width=7, bd=5)
oilEntry.grid(row=1, column=1, pady=9, padx=10)
oilEntry.insert(0, 0)
oilEntry.bind("<FocusIn>", clear_initial_zero)
oilEntry.bind("<FocusOut>", restore_initial_zero)

dalLabel=Label(groceryFrame, text='Dal  (₹150)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
dalLabel.grid(row=2, column=0, pady=9, padx=10, sticky='w')

dalEntry=Entry(groceryFrame, font=('times new roman',15,'bold'), width=7, bd=5)
dalEntry.grid(row=2, column=1, pady=9, padx=10)
dalEntry.insert(0, 0)
dalEntry.bind("<FocusIn>", clear_initial_zero)
dalEntry.bind("<FocusOut>", restore_initial_zero)

wheatLabel=Label(groceryFrame, text='Wheat  (₹40)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
wheatLabel.grid(row=3, column=0, pady=9, padx=10, sticky='w')

wheatEntry=Entry(groceryFrame, font=('times new roman',15,'bold'), width=7, bd=5)
wheatEntry.grid(row=3, column=1, pady=9, padx=10)
wheatEntry.insert(0, 0)
wheatEntry.bind("<FocusIn>", clear_initial_zero)
wheatEntry.bind("<FocusOut>", restore_initial_zero)

sugarLabel=Label(groceryFrame, text='Sugar  (₹40)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
sugarLabel.grid(row=4, column=0, pady=9, padx=10, sticky='w')

sugarEntry=Entry(groceryFrame, font=('times new roman',15,'bold'), width=7, bd=5)
sugarEntry.grid(row=4, column=1, pady=9, padx=10)
sugarEntry.insert(0, 0)
sugarEntry.bind("<FocusIn>", clear_initial_zero)
sugarEntry.bind("<FocusOut>", restore_initial_zero)

teaLabel=Label(groceryFrame, text='Tea  (₹300)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
teaLabel.grid(row=5, column=0, pady=9, padx=10, sticky='w')

teaEntry=Entry(groceryFrame, font=('times new roman',15,'bold'), width=7, bd=5)
teaEntry.grid(row=5, column=1, pady=9, padx=10)
teaEntry.insert(0, 0)
teaEntry.bind("<FocusIn>", clear_initial_zero)
teaEntry.bind("<FocusOut>", restore_initial_zero)

#Stationary Frame
statFrame=LabelFrame(productFrame, text='Stationary', font=('times new roman',15,'bold'), fg='gold', bd=8, relief=GROOVE, bg='gray20')
statFrame.grid(row=0, column=3, sticky="nsew")

envelopeLabel=Label(statFrame, text='Envelope  (₹10)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
envelopeLabel.grid(row=0, column=0, pady=9, padx=10, sticky='w')

envelopeEntry=Entry(statFrame, font=('times new roman',15,'bold'), width=7, bd=5)
envelopeEntry.grid(row=0, column=2, pady=9, padx=10)
envelopeEntry.insert(0, 0)
envelopeEntry.bind("<FocusIn>", clear_initial_zero)
envelopeEntry.bind("<FocusOut>", restore_initial_zero)

penLabel=Label(statFrame, text='Pen  (₹10)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
penLabel.grid(row=1, column=0, pady=9, padx=10, sticky='w')

penEntry=Entry(statFrame, font=('times new roman',15,'bold'), width=7, bd=5)
penEntry.grid(row=1, column=2, pady=9, padx=10)
penEntry.insert(0, 0)
penEntry.bind("<FocusIn>", clear_initial_zero)
penEntry.bind("<FocusOut>", restore_initial_zero)

calciLabel=Label(statFrame, text='Calculator  (₹1000)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
calciLabel.grid(row=2, column=0, pady=9, padx=10, sticky='w')

calciEntry=Entry(statFrame, font=('times new roman',15,'bold'), width=7, bd=5)
calciEntry.grid(row=2, column=2, pady=9, padx=10)
calciEntry.insert(0, 0)
calciEntry.bind("<FocusIn>", clear_initial_zero)
calciEntry.bind("<FocusOut>", restore_initial_zero)

notebookLabel=Label(statFrame, text='Notebook  (₹100)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
notebookLabel.grid(row=3, column=0, pady=9, padx=10, sticky='w')

notebookEntry=Entry(statFrame, font=('times new roman',15,'bold'), width=7, bd=5)
notebookEntry.grid(row=3, column=2, pady=9, padx=10)
notebookEntry.insert(0, 0)
notebookEntry.bind("<FocusIn>", clear_initial_zero)
notebookEntry.bind("<FocusOut>", restore_initial_zero)

staplerLabel=Label(statFrame, text='Stapler  (₹100)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
staplerLabel.grid(row=4, column=0, pady=9, padx=10, sticky='w')

staplerEntry=Entry(statFrame, font=('times new roman',15,'bold'), width=7, bd=5)
staplerEntry.grid(row=4, column=2, pady=9, padx=10)
staplerEntry.insert(0, 0)
staplerEntry.bind("<FocusIn>", clear_initial_zero)
staplerEntry.bind("<FocusOut>", restore_initial_zero)

glueLabel=Label(statFrame, text='Glue  (₹10)', font=('times new roman',15,'bold'), bg='gray20', fg='white')
glueLabel.grid(row=5, column=0, pady=9, padx=10, sticky='w')

glueEntry=Entry(statFrame, font=('times new roman',15,'bold'), width=7, bd=5)
glueEntry.grid(row=5, column=2, pady=9, padx=10)
glueEntry.insert(0, 0)
glueEntry.bind("<FocusIn>", clear_initial_zero)
glueEntry.bind("<FocusOut>", restore_initial_zero)

#-----------------------------------------------------------------------------------------------------------------------

# Create the bill area frame (right side)
billFrame = Frame(productFrame, bd=8, relief=GROOVE, width=400)
billFrame.grid(row=0,column=4, padx=160, sticky="nsew")

billareaLabel=Label(billFrame, text='Bill Area', font=('times new roman',15,'bold'),bd=8, relief=GROOVE, padx=185)
billareaLabel.pack()

scrollbar=Scrollbar(billFrame, orient='vertical')
scrollbar.pack(side=RIGHT, fill=Y)

textarea=Text(billFrame, height=18, width=55, yscrollcommand=scrollbar.set)
textarea.pack()
scrollbar.config(command=textarea.yview)

billmenuFrame=LabelFrame(root, text='Bill Menu', font=('times new roman',15,'bold'), fg='gold', bd=8, relief=GROOVE, bg='gray20')
billmenuFrame.pack(fill=Y)

cosmeticpriceLabel=Label(billmenuFrame, text='Cosmetics Price', font=('times new roman',15,'bold'), fg='white', bg='gray20')
cosmeticpriceLabel.grid(row=0, column=0, pady=9, padx=10, sticky='w')

cosmeticpriceEntry=Entry(billmenuFrame, font=('times new roman', 15, 'bold'), bd=5, width=10)
cosmeticpriceEntry.grid(row=0, column=1, pady=9, padx=10)

grocerypriceLabel=Label(billmenuFrame, text='Grocery Price', font=('times new roman',15,'bold'), fg='white', bg='gray20')
grocerypriceLabel.grid(row=1, column=0, pady=9, padx=10, sticky='w')

grocerypriceEntry=Entry(billmenuFrame, font=('times new roman', 15, 'bold'), bd=5, width=10)
grocerypriceEntry.grid(row=1, column=1, pady=9, padx=10)

statpriceLabel=Label(billmenuFrame, text='Stationary Price', font=('times new roman',15,'bold'), fg='white', bg='gray20')
statpriceLabel.grid(row=3, column=0, pady=9, padx=10, sticky='w')

statpriceEntry=Entry(billmenuFrame, font=('times new roman', 15, 'bold'), bd=5, width=10)
statpriceEntry.grid(row=3, column=1, pady=9, padx=10)

#-----------------------------------------------------------------------------------------------------------------------

cosmetictaxLabel=Label(billmenuFrame, text='Cosmetics Tax (12%)', font=('times new roman',15,'bold'), fg='white', bg='gray20')
cosmetictaxLabel.grid(row=0, column=2, pady=9, padx=10, sticky='w')

cosmetictaxEntry=Entry(billmenuFrame, font=('times new roman', 15, 'bold'), bd=5, width=10)
cosmetictaxEntry.grid(row=0, column=3, pady=9, padx=10)

grocerytaxLabel=Label(billmenuFrame, text='Grocery Tax (5%)', font=('times new roman',15,'bold'), fg='white', bg='gray20')
grocerytaxLabel.grid(row=1, column=2, pady=9, padx=10, sticky='w')

grocerytaxEntry=Entry(billmenuFrame, font=('times new roman', 15, 'bold'), bd=5, width=10)
grocerytaxEntry.grid(row=1, column=3, pady=9, padx=10)

#drinkstaxLabel=Label(billmenuFrame, text='Drinks Tax', font=('times new roman',15,'bold'), fg='white', bg='gray20')
#drinkstaxLabel.grid(row=2, column=2, pady=9, padx=10, sticky='w')

#drinkstaxEntry=Entry(billmenuFrame, font=('times new roman', 15, 'bold'), bd=5, width=10)
#drinkstaxEntry.grid(row=2, column=3, pady=9, padx=10)

stattaxLabel=Label(billmenuFrame, text='Stationary Tax (15%)', font=('times new roman',15,'bold'), fg='white', bg='gray20')
stattaxLabel.grid(row=3, column=2, pady=9, padx=10, sticky='w')

stattaxEntry=Entry(billmenuFrame, font=('times new roman', 15, 'bold'), bd=5, width=10)
stattaxEntry.grid(row=3, column=3, pady=9, padx=10)

#Button Frame
buttonFrame=Frame(billmenuFrame, bd=8, relief=GROOVE)
buttonFrame.grid(row=0, column=7, rowspan=6, padx=70)

#Total Button
totalButton=Button(buttonFrame,text='Total',font=('arial',16,'bold'),bg='gray20',fg='white',bd=5,width=8,pady=10, command=total, cursor='hand2')
totalButton.grid(row=0, column=0, pady=20, padx=5)

#Bill Button
billButton=Button(buttonFrame,text='Bill',font=('arial',16,'bold'),bg='gray20',fg='white',bd=5,width=8, pady=10, cursor='hand2', command=bill_area)
billButton.grid(row=0, column=1, pady=20, padx=5)

#Email Button
emailButton=Button(buttonFrame,text='Email',font=('arial',16,'bold'),bg='gray20',fg='white',bd=5,width=8,pady=10, cursor='hand2', command=send_email)
emailButton.grid(row=0, column=2, pady=20, padx=5)

#Print Button
printButton=Button(buttonFrame,text='Print',font=('arial',16,'bold'),bg='gray20',fg='white',bd=5,width=8,pady=10, cursor='hand2', command=print_bill)
printButton.grid(row=0, column=3, pady=20, padx=5)

#Clear Button
clearButton=Button(buttonFrame,text='Clear',font=('arial',16,'bold'),bg='gray20',fg='white',bd=5,width=8,pady=10,  cursor='hand2', command=clear)
clearButton.grid(row=0, column=4, pady=20, padx=5)

excelButton=Button(buttonFrame, text='Excel', font=('arial',16,'bold'),bg='gray20',fg='white',bd=5,width=8,pady=10,  cursor='hand2', command=store_customer_details)
excelButton.grid(row=0, column=5, pady=20, padx=5)

root.mainloop()
